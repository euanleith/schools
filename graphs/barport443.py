from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

import sys,os
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import funcs

# graphs

def stackedBars(bottomBarValues, topBarValues, xlabels, ylabel, \
               bottomBarLabel = 'true', topBarLabel = 'false', \
               bottomBarColour = (0, 1, 0, 0.5513898339486097), \
               topBarColour = (1, 0, 0, 0.4675564992027787), \
               edgeColour = (0,0,0,0.5), total = -1, title = '', \
               width = 0.35, gap = 0.1):

    totalWidth=np.arange(len(bottomBarValues),dtype='float')#change this with more years
    for i in range(len(totalWidth)):
        totalWidth[i]*=width+gap
    # get np.arrays for all bool types
    t=np.array(bottomBarValues,dtype='float')
    f=np.array(topBarValues,dtype='float')

    # convert to proportions (for labels)
    for i in range(max(len(t),len(f))):
        s=t[i]+f[i]
        if s != 0:
            t[i]/=s
            f[i]/=s

    rectsTrue = plt.bar(totalWidth, t, width,
                        color=bottomBarColour,
                        edgecolor=edgeColour,
                        label=bottomBarLabel)
    rectsFalse = plt.bar(totalWidth, f, width,
                        color=topBarColour,
                        edgecolor=edgeColour,
                        bottom=t,
                        label=topBarLabel)

    
    if total != -1:
        txt=[] # label for proportion of all schools that had successful zgrab
        for i in range(len(bottomBarValues)):
            numFound=bottomBarValues[i]+topBarValues[i]
            txt.append(str(numFound)+'/'+str(total[i]))

        funcs.labelBars(rectsFalse,txt=txt,y=1)
        
    #labelBars(rectsTrue)
    #labelBars(rectsFalse,txt=[title]*len(bottomBarValues),y=0)#~~
    #plt.xticks([x+(len(years)-1)*(width/2) for x in range(len(bottomBarValues))],boolNames)
    plt.xticks(totalWidth,xlabels)
    plt.ylabel(ylabel)

    #totalWidth=totalWidth+width

    plt.title(title)
    plt.yticks=(np.arange(0,1,0.1))

# data 

years=["schools2018.xlsx","schools2019.xlsx"]

just80=[]
ssl=[]
for year in years:
    just80.append(0)
    ssl.append(0)

total=[]
for year in years:
    wb = load_workbook('../'+year)
    ws = wb.active

    statusCodes=[ws['P'],ws['R'],ws['AA']]
    errors=ws['AL']
    if year != 'schools2018.xlsx': #...
        total.append(len(statusCodes[0]))
    else:
        total.append(716)

    for i in range(len(statusCodes[0])):
        p80=statusCodes[0][i].value
        if p80 and p80 != -1 and p80 != 999:
            error=False
            if year != 2018:
                # check if error implied a 443 listener
                # form: port|error$port|error 
                if errors[i].value and '$' in errors[i].value:
                    ports = errors[i].value.split('$')
                    # those errors which when returned on 443 imply a 443 listener
                    errortypes=['connection reset by peer','oversized record']
                    #connection refused?
                    #list isnt right...
                    for port in ports[:len(ports)-2]:
                        e = port.split('|')
                        if e[0] != '80':
                            for errortype in errortypes:
                                if errortype in e[1]:
                                    error=True
                                    break
                        
                #error = errors[i].value and 'i/o timeout' in errors[i].value
                #need to know if this was from 443 too......
                error=False
            p443with=statusCodes[2][i].value
            p443without=statusCodes[1][i].value
            if (not p443with or p443with == -1 or p443with == 999) and \
               (not p443without or p443without == -1 or p443without == 999) and \
               not error:
                just80[years.index(year)]+=1
            else:
                ssl[years.index(year)]+=1

                    
    years[years.index(year)]=year.split('schools')[1].split('.')[0]

plt.figure(figsize=(5,7))

for i in range(len(years)):
    numFound=ssl[i]+just80[i]
    years[i]+=' ('+str(numFound)+'/'+str(total[i])+')'
    
stackedBars(ssl, just80, years, 'Proportion of schools', \
            #title='schools with https', \
            width=0.05, gap = 0.01)

plt.savefig("../imgs/port443.png")
plt.show()

    
