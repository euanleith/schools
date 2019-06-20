from openpyxl import load_workbook
import matplotlib.pyplot as plt

import sys,os
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import funcs

MIN_KEY_DUPS = 0

ports=['443 without sni','443 with sni']
years=["schools2018.xlsx","schools2019.xlsx"]
colourDct={}
#plt.figure(figsize=(5,5))
for year in years:
    plt.subplot(1,2,years.index(year)+1)
    wb = load_workbook('../'+year)
    ws = wb.active
    year=year.split('schools')[1].split('.')[0]

    fp=[ws['U'],ws['AD']]
    urls=ws['AM']
    
    names=[ws['T'],ws['AC']]

    

    for port in ports:
        dups=dict()
        for row in fp[ports.index(port)]:
            if row.value:
               funcs.addDups(dups,str(row.value),[fp[ports.index(port)].index(row)])
        dups=funcs.sortDict(dups,minDups=MIN_KEY_DUPS) # keep keys associated with multiple schools

        
##        print('\n\n'+year,port)
##        for key in dups:
##            # check to see if there's any correlation between these urls which share keys
##            namesDct=dict()
##            for row in dups[key]:
##                funcs.addDups(namesDct,names[ports.index(port)][row].value)
##            namesDct=funcs.sort(namesDct)
##            print('\n\n-'+key+':',namesDct)
    
    # each key and how many people are using that key
    # want to split by number of keys shared
    dct={}
    temp={}
    for key in dups:
        dct=funcs.addDups(dct,len(dups[key]),len(dups[key]))
        colourDct=funcs.mapColour(colourDct,len(dups[key]))
    dct=funcs.sortDict(dct,reverse=True)
    
    values=list(dct.values())
    #make 1 = 'unique'?
    keys=list(dct.keys())

    colours=[]
    for item in keys:
        colours.append(colourDct[item])
    #colours.reverse()
    
    labels = keys
        
    patches,texts= plt.pie(values,
                            #autopct='%1.1f%%', #need to inc junk =
                            #labels=labels,
                            colors=colours,
                            wedgeprops={'linewidth': 0.5,"edgecolor":(0,0,0,0.5)},  
                            startangle=0)#140
    if year == 2018:
        total=716
    else:
        total = len(fp[0])
    #plt.text(1,1,str(total-sum(values))+'/'+str(total))
    #colour should be based on the key
    #-would be nice if it used the first key for each group

    plt.axis('equal')
    plt.legend(patches,labels,loc='lower right')
    plt.title(year+' ('+str(sum(values))+'/'+str(total)+')')

plt.subplots_adjust(wspace=0.5)
#plt.suptitle('Duplicate keys')
plt.tight_layout()
#plt.savefig('../imgs/fp.png')
plt.show()

        
