from openpyxl import load_workbook
import pandas as pd
import geopandas
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

import sys,os
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import funcs

years=["schools2018.xlsx","schools2019.xlsx"]
recordtypes=['AAAA', 'DS', 'CAA', 'SPF']
cols=['J','L','M','N']

colours=['coral','teal','gold','yellowgreen']
handles=[]
for i in range(len(recordtypes)):
    #colours.append(np.random.uniform(0,1,size=3))
    handles.append(mpatches.Patch(color=colours[i],label=recordtypes[i])) # for legend

# background image (ireland)
base = geopandas.read_file('Ireland.GeoJSON')
    
for year in years:
    wb = load_workbook('../'+year)
    ws = wb.active
    year=year.split('schools')[1].split('.')[0]

    if year == '2018': # if coords already
        #better if statement?
        latCol=list(ws['D'])
        longCol=list(ws['E'])
        for i in range(min(len(latCol),len(longCol))):
            latCol[i]=latCol[i].value
            longCol[i]=longCol[i].value
    else: # if eircodes
        eircodes=[row.value for row in ws['E']]
        latCol, longCol = funcs.getCoordsFromPostcodes(eircodes)

    coordsByType=[] # list of coords for each recordtype
    # get all coords for a given func
    for i in range(len(cols)):
        col=ws[cols[i]]
        lats=[]
        longs=[]
        for row in range(len(latCol)):
            lat=latCol[row]
            long=longCol[row]
            item=col[row].value
            if lat and long and item and\
               float(lat) < 100 and float(long) < 100:
                if 'no' not in str(item):#func here
                    lats.append(lat)
                    longs.append(long)
        coordsByType.append([lats,longs])

    # plot base
    ax = base.plot(color='white', edgecolor='black')

    # mapping multiple types of points onto the same base
    for i in range(len(coordsByType)):
        coords=coordsByType[i]
        funcs.mapCoords(coords[0], coords[1], ax, colour=colours[i])

    plt.legend(handles=handles)
    plt.suptitle(year)
    
plt.show()
