from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

plt.figure(figsize=(14,5))
#colours={'hasnt':(1,0,0,np.random.uniform(0.2,1)),'has':(0,1,0,np.random.uniform(0.2,1))}
colours={'hasnt': (1, 0, 0, 0.4675564992027787), 'has': (0, 1, 0, 0.5513898339486097)}

years=["schools2018.xlsx","schools2019.xlsx"]
for year in years:
    #plt.figure(years.index(year)+1)
    plt.subplot(1,2,years.index(year)+1)
    wb = load_workbook('../'+year)
    ws = wb.active
##    if year != '2018':
##        wbdups = load_workbook('dups.xlsx')
##        wsdups = wbdups.active
    
    year=year.split('schools')[1].split('.')[0]
    
##    print('\n['+year+']: ('+str(len(ws['J']))+')')

    recordtypes=['AAAA', 'MX', 'DS', 'CAA', 'SPF', 'TXT','CNAME','A']
    recordtypesdict = dict({'AAAA':[], 'MX':[], 'DS':[], 'CAA':[], \
                            'SPF':[], 'TXT':[],'CNAME':[],'A':[]})
    cols=['J','K','L','M','N','O','AJ','AK']

    items=dict({'AAAA':[], 'MX':[], 'DS':[], 'CAA':[], \
                'SPF':[], 'TXT':[],'CNAME':[],'A':[]})#name(hasnt)
    
    urls=list(ws['AM'])

    for recordtype in range(len(recordtypes)):
        col=list(ws[cols[recordtype]])
        
        # inc dups
        #if year != '2018':
        #    col.extend(wsdups[cols[recordtype]])
        for row in range(len(col)):
            hasItem=False # if checked for a given DNS item that year
            
            if col[row].value:
                hasItem=True
                if not 'no ' + recordtypes[recordtype] in col[row].value:
                    recordtypesdict[recordtypes[recordtype]].append(urls[row].value)
                else:
                    items[recordtypes[recordtype]].append(urls[row].value)

        if not hasItem:
            items.pop(recordtypes[recordtype])
                    
##    for key in recordtypesdict:
##       print(key+':',len(recordtypesdict[key]))
##       if not 'no' + recordtypes[recordtype] in key:# and key != 'AAAA':
##            print(recordtypesdict[key])

    #display

#~
    hasnt=[]
    for value in list(items.values()):
        hasnt.append(len(value))
    N = len(hasnt)

    has=[]
    cnt=0
    for recordtype in recordtypesdict:
        if cnt >= len(hasnt):
            break
        has.append(len(recordtypesdict[recordtype]))
        cnt+=1
#~
    ind = np.arange(N)    # the x locations for the groups
    width = 0.35       # the width of the bars: can also be len(x) sequence

    p1 = plt.bar(ind, has, width,
                 color=colours['has'],
                edgecolor=(0,0,0,0.5))
    p2 = plt.bar(ind, hasnt, width,
                 bottom=has,
                 color=colours['hasnt'],
                 edgecolor=(0,0,0,0.5))

    plt.ylabel('Number of schools')
    if year == '2018':
        total=716#~
    else:
        total=len(urls)
    #print('('+str(has[0]+hasnt[0])+'/' + str(total)+')')
    plt.title(year+' ('+str(has[0]+hasnt[0])+'/'+str(total)+')')
    plt.xticks(ind, recordtypes)
    plt.yticks(np.arange(0, has[0]+hasnt[0], 30))
    
#plt.legend((p1[0], p2[0]), ('has', 'hasnt'),bbox_to_anchor=(1.2,1))
#plt.suptitle('Schools dns')
plt.savefig('../imgs/DNS.png')
plt.show()

                
