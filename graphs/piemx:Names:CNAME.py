from openpyxl import load_workbook
import matplotlib.pyplot as plt

import sys,os
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import funcs

MIN_DUPS=10 # minimum duplicates needed to be shown on graph, otherwise included in 'other'

#cols=['K','T','AC','AJ']
cols=['K','AJ']
names=['MX','CNAME']
#names=['MX','names 443 without sni','names 443 with sni','CNAME']
allHosts=dict()

for i in range(len(cols)):
    plt.figure(i+1)
    years=['schools2018.xlsx','schools2019.xlsx']
    for year in years:
        wb = load_workbook('../'+year)
        ws = wb.active

        info=ws[cols[i]]
        urls=ws['H']
        
        hosts=dict()
        h={'google','outlook','blacknight','hosts',\
               'reg365','irishdomains','cloudaccess',\
               'tstechnology','mailspamprotection','elive',\
               'host','scoilnet','spamtitan','spamdoctor',\
               'forwardinfo','messagelabs','novara',\
               'odin','plesk','weebly','parallels panel',\
               'squarespace','avenir','carrigtwohill',\
               'stdeclanscollege','maryimmaculate','schoolwiselearning',\
               'dnsireland','schoolwiselearning','sgvps','lurtel',\
               'dmacmedia','cloudflare','finalsite','stackpathdns',\
               'greenschoolsonline'}#~

        numFailed=0
        for row in range(len(info)):
            if info[row].value:
                if 'no ' + names[i] not in info[row].value:
                    found=False
                    for host in h:
                        if host in info[row].value.lower():
                            hosts=funcs.addDups(hosts,host)
                            allHosts=funcs.mapColour(allHosts,host)
                            found=True
                            break
                    if not found:
                        if urls[row].value and urls[row].value.split('.')[0] not in info[row].value:
                            hosts=funcs.addDups(hosts,info[row].value)
                            allHosts=funcs.mapColour(allHosts,info[row].value)
                        else:
                            hosts=funcs.addDups(hosts,'self-hosted')
                            allHosts=funcs.mapColour(allHosts,'self-hosted')
            else:
                numFailed+=1

        hosts=funcs.sortDict(hosts,MIN_DUPS, reverse=True, showOther=True)

        if names[i] != 'CNAME':
            plt.subplot(len(years),1,years.index(year)+1)
        year=year.split('schools')[1].split('.')[0]
        if year == '2018':
            total=716#~
        else:
            total=len(urls)
        plt.title(year+' ('+str(len(urls)-numFailed)+'/'+str(total)+')',y=0.8)#~

        labels = list(hosts.keys())

        colours=[]
        for host in hosts:
            if host == 'other':#~
                colours.append('coral')
            else:
                colours.append(allHosts[host])
        
        patches,texts= plt.pie(hosts.values(),
                                #autopct='%1.1f%%', #need to inc junk =
                                #labels=labels,
                                colors=colours,
                                wedgeprops={'linewidth': 0.5,"edgecolor":(0,0,0,0.5)},
                                startangle=0)#140
        #plt.text(0.5,0.5,str(len(urls)-numFailed)+'/'+str(len(urls)))

        plt.axis('equal')
        plt.legend(patches,labels,loc='lower right')
    
    plt.subplots_adjust(wspace=0.5)
    #plt.suptitle(names[i])
    plt.tight_layout()
    #plt.savefig('../imgs/'+names[i]+'.png')
    
plt.show()

    
