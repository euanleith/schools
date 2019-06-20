from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

import sys,os
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import funcs

# getting data
    
ports=['443 with SNI','443 without SNI']
years=["schools2018.xlsx","schools2019.xlsx"]
boolNames=['Timely','Not self signed','Browser trusted']
totalLengths=[] # total num schools for each year

plt.figure(figsize=(14,5))

for port in range(len(ports)):
    #plt.figure(port+1)
    plt.subplot(1,2,port+1)
    true=dict()
    false=dict()
    for year in years:
        true.update({year:dict()})
        false.update({year:dict()})
        wb = load_workbook('../'+year)
        ws = wb.active

        timely=[ws['AE'],ws['V']]
        selfSigned=[ws['AF'],ws['W']]
        browserTrusted=[ws['AI'],ws['Z']]
        bools=[timely,selfSigned,browserTrusted]
        totalLengths.append(len(timely[0]))
        

        for b in range(len(bools)):
            for row in range(len(bools[b][port])):
                item=str(bools[b][port][row].value)
                if item:
                    if not 'not' in boolNames[b].lower():
                        if item.lower() == 'true':
                            true[year]=funcs.addDups(true[year],boolNames[b])
                        elif item.lower() == 'false':
                            false[year]=funcs.addDups(false[year],boolNames[b])
                    else:
                        if item.lower() == 'true':
                            false[year]=funcs.addDups(false[year],boolNames[b])
                        elif item.lower() == 'false':
                            true[year]=funcs.addDups(true[year],boolNames[b])

# graphs
                            
    width = 0.35
            
    totalWidth=np.arange(len(true[years[0]]))#change this with more years
    for year in years:
        # get np.arrays for all bool types
        t=np.array(list(true[year].values()),dtype='float')
        f=np.array(list(false[year].values()),dtype='float')
        
        # convert to proportions (for labels)
        for i in range(max(len(t),len(f))):
            s=t[i]+f[i]
            t[i]/=s
            f[i]/=s

        rectsTrue = plt.bar(totalWidth, t, width,
                         #color='#0000DD',
                        color=(0, 1, 0, 0.5513898339486097),
                         edgecolor=(0,0,0,0.5),
                         label='true')
        rectsFalse = plt.bar(totalWidth, f, width,
                         #color='#ED0020',
                        color=(1, 0, 0, 0.4675564992027787),
                         edgecolor=(0,0,0,0.5),
                         bottom=t,
                         label='false')

#~
        txt=[] # label for proportion of all schools that had successful zgrab
        for i in range(len(true[year].values())):
            numFound=list(true[year].values())[i]+list(false[year].values())[i]
            if year == 'schools2018.xlsx':
                total=716#~
            else:
                total=totalLengths[years.index(year)]
            txt.append(str(numFound)+'/'+str(total))
        funcs.labelBars(rectsFalse,txt=txt,y=1)
        
        funcs.labelBars(rectsTrue,txt=[year.split('schools')[1].split('.')[0]]*len(boolNames),y=0)#~~
        plt.xticks([x+(len(years)-1)*(width/2) for x in range(len(boolNames))],boolNames)
        
        totalWidth=totalWidth+width


    plt.ylabel('Proportion of schools')
    plt.title(ports[port])
    plt.yticks=(np.arange(0,1,0.1))

#plt.legend((rectsTrue[0],rectsFalse[0]),('True','False'),bbox_to_anchor=(1.2,1))
#plt.suptitle('bools')
#plt.savefig('../imgs/barbools.png')
plt.show()

