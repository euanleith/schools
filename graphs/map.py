from openpyxl import load_workbook
import pandas as pd
import geopandas
import matplotlib.pyplot as plt

import sys,os
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import funcs

years=["schools2018.xlsx","schools2019.xlsx"]
base = geopandas.read_file('Ireland.GeoJSON')

for year in years:
    wb = load_workbook('../'+year)
    ws = wb.active
    year=year.split('schools')[1].split('.')[0]

    if year == '2018': # if coords already
        #better if statement?
        latCol=list(ws['D'])
        longCol=list(ws['E'])
        for i in range(min(len(latCol),len(longCol))):
            latCol[i]=latCol[i].value
            longCol[i]=longCol[i].value
    else: # if eircodes
        eircodes=[row.value for row in ws['E']]
        latCol, longCol = funcs.getCoordsFromPostcodes(eircodes)

    lats=[]
    longs=[]
    for i in range(min(len(latCol),len(longCol))):
        lat=latCol[i]
        long=longCol[i]
        if lat and long and\
           lat < 90 and lat > -90 and\
           long < 180 and long > -180: #weird outliers...
            lats.append(lat)
            longs.append(long)

    ax=base.plot(color='white', edgecolor='black')
    funcs.mapCoords(lats, longs, ax)
    plt.suptitle(year)

plt.show()
