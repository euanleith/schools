#shows t and f in different colours
from openpyxl import load_workbook
import pandas as pd
import geopandas
import matplotlib.pyplot as plt

import sys,os
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import funcs

wb = load_workbook('../schools2018.xlsx')
ws = wb.active

latCol=ws['D']
longCol=ws['E']
col=ws['Z']

flats=[]
flongs=[]
tlats=[]
tlongs=[]
for i in range(len(latCol)):
    lat=latCol[i].value
    long=longCol[i].value
    item=str(col[i].value)
    if lat and long and item and\
       float(lat) < 100 and float(long) < 100:
        if item.lower() == 'false':
            flats.append(lat)
            flongs.append(long)
        elif item.lower() == 'true':
            tlats.append(lat)
            tlongs.append(long)
      
world = geopandas.read_file('../Ireland.GeoJSON')
ax=world.plot(color='white', edgecolor='black')
funcs.mapCoords(flats, flongs, ax)
funcs.mapCoords(tlats, tlongs, ax, colour='b')

plt.show()
