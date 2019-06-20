from openpyxl import load_workbook
import matplotlib.pyplot as plt
import operator
import numpy as np
import pandas as pd

import sys,os
sys.path.insert(1, os.path.join(sys.path[0], '..'))
from funcs import mapColour
from funcs import addDups
from funcs import sortDict

MIN_DUPS=0 # minimum duplicates needed to be shown on graph, otherwise included in 'other'

# read from iana
df = pd.read_csv('../tls-parameters-4.csv')

values = df['Value']
description = df['Description']
names=[]

# ech data, need to fix names for each value
# if finds '-', split into 2 nums on either side, then sub these in n, then add current name n times  
for i in range(len(values)):
    value=values[i]
    if '-' in value:
        nums = value.split('-')
        if '*' in value:
            nums[0]=nums[0]+'0x00'
            nums[1]='0x'+nums[1].split('*')[0]+'0xff'
        if not '0x' in nums[1]:
            nums[1]=nums[0].split(',')[0]+',0x'+nums[1]
        for num in range(len(nums)):
            t=nums[num].split('0x')
            nums[num]=t[1].split(',')[0]+t[2]
        n = int(nums[1],16)-int(nums[0],16)
        for j in range(n):
            names.append(description[i])
    else:
        names.append(description[i])
        
recommendations = df['Recommended']
badSuites=[]
for i in range(len(recommendations)):
    if recommendations[i] == 'N':
        badSuites.append(names[i])


cols=['Y','AH']
ports=['443 without SNI','443 with SNI']
allSuites=dict()#messy~

for i in range(len(cols)):
    plt.figure(i+1)
    years=['schools2018.xlsx','schools2019.xlsx']
    for year in years:
        wb = load_workbook('../'+year)
        ws = wb.active

        info=ws[cols[i]]
        urls=ws['H']
        
        suites=dict()

        numFailed=0
        for row in range(len(info)):
            if info[row].value and info[row].value != -1:
                suites=addDups(suites,names[info[row].value])
                if names[info[row].value] in badSuites:
                    colour = (1,0,0,np.random.uniform(0.2,0.8))
                else:
                    colour = (0,1,0,np.random.uniform(0.2,0.8))
                allSuites=mapColour(allSuites,names[info[row].value],colour)
            else:
                numFailed+=1

        suites=sortDict(suites,minDups=MIN_DUPS,reverse=True)

        plt.subplot(2,1,years.index(year)+1)
        year=year.split('schools')[1].split('.')[0]
        if year == '2018':
            total=716#~
        else:
            total=len(urls)
        plt.title(year+' ('+str(len(urls)-numFailed)+'/'+str(total)+')',y=0.8)#~

        labels = suites.keys()

        colours=[]
        for suite in suites:
            colours.append(allSuites[suite])
        
        patches,texts= plt.pie(suites.values(),
                                #autopct='%1.1f%%', #need to inc junk =
                                #labels=labels,
                                colors=colours,
                                wedgeprops={'linewidth': 0.5,"edgecolor":(0,0,0,0.5)},
                                startangle=0)#140
        #plt.text(0.5,0.5,str(len(urls)-numFailed)+'/'+str(len(urls)))

        good=0
        bad=0
        for suite in allSuites:
            if suite in suites.keys():
                if allSuites[suite][0]==1:
                    bad+=suites[suite]
                else:
                    good+=suites[suite]

        #maybe do bar instead of this, and only show with sni
        plt.text(2,0.3,str(round(good/(good+bad)*100,2))+'% returned recommended suites')
        

        plt.axis('equal')
        plt.legend(patches,labels,loc='best',bbox_to_anchor=(1,0.5)) #should only need one legend

    plt.subplots_adjust(wspace=0.5)
    plt.suptitle(ports[i])
    plt.tight_layout()
    #plt.savefig('../imgs/cipher '+ports[i]+'.png')
    
plt.show()

    
