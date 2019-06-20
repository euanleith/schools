# 5
import csv
import dns.resolver #import the module

import os, sys, argparse, tempfile, gc
import json
import jsonpickle # install via  "$ sudo pip install -U jsonpickle"
import time, datetime
from dateutil import parser as dparser  # for parsing time from comand line and certs
import pytz # for adding back TZ info to allow comparisons
import subprocess
import traceback

#from https://github.com/sftcd/surveys ...
# locally in $HOME/code/surveys
#def_surveydir=os.environ['HOME']+'/code/surveys'#~
#sys.path.insert(0,def_surveydir)#~
#import FreshGrab.py

#codedir=os.path.dirname(os.path.realpath(__file__))
#pdir=os.path.dirname(codedir)
#sys.path.insert(0,pdir)

sys.path.insert(1, os.path.join(sys.path[0], '..'))
from SurveyFuncs import *

from openpyxl import load_workbook

wb = load_workbook("../schools4.xlsx")
ws = wb.active


myResolver = dns.resolver.Resolver() #create a new instance named 'myResolver'

# command line arg handling 
argparser=argparse.ArgumentParser(description='Read a CSV, one (default last) column of which is a URL, then append a bunch of Web derived values to the row and write out to the output file')
argparser.add_argument('-i','--infile',     
                    dest='infile',
                    help='CSV file containing list of domains')
argparser.add_argument('-o','--output_file',     
                    dest='outfile',
                    help='CSV file in which to put records (one per line)')
argparser.add_argument('-c','--col',     
                    dest='col',
                    help='column from input file that has the URL')
argparser.add_argument('-s','--scandate',     
                    dest='scandatestring',
                    help='time at which to evaluate certificate validity')
args=argparser.parse_args()

# scandate is needed to determine certificate validity, so we support
# the option to now use "now"
if args.scandatestring is None:
    scandate=datetime.datetime.utcnow().replace(tzinfo=pytz.UTC)
    #print >> sys.stderr, "No (or bad) scan time provided, using 'now'"
else:
    scandate=dparser.parse(args.scandatestring).replace(tzinfo=pytz.UTC)
    print ("Scandate: using " + args.scandatestring + "\n", file=sys.stderr)


def_country='IE'
country=def_country

def wr_dummy(row):#~def wr_dummy(row,wr)
    dummy=[]
    # port 80
    # analysis[port]['status_code']=-1
    dummy.append(-1)
    # analysis[port]['body_len']=-1
    dummy.append(-1)
    # port 443, no SNI
    # analysis[port]['status_code']=-1
    dummy.append(-1)
    # analysis[port]['body_len']=-1
    dummy.append(-1)
    # analysis[port]['names']=''
    dummy.append('')
    # analysis[port]['fp']=''
    dummy.append('')
    # analysis[port]['timely']=-1
    dummy.append(-1)
    # analysis[port]['self_signed']=-1
    dummy.append(-1)
    # analysis[port]['rsalen']=-1
    dummy.append(-1)
    # analysis[port]['cipher_suite']=-1
    dummy.append(-1)
    # analysis[port]['browser_trusted']=-1
    dummy.append(-1)
    # port 443, with SNI
    # analysis[port]['status_code']=-1
    dummy.append(-1)
    # analysis[port]['body_len']=-1
    dummy.append(-1)
    # analysis[port]['names']=''
    dummy.append('')
    # analysis[port]['fp']=''
    dummy.append('')
    # analysis[port]['timely']=-1
    dummy.append(-1)
    # analysis[port]['self_signed']=-1
    dummy.append(-1)
    # analysis[port]['rsalen']=-1
    dummy.append(-1)
    # analysis[port]['cipher_suite']=-1
    dummy.append(-1)
    # analysis[port]['browser_trusted']=-1
    dummy.append(-1)
    for i in range(len(dummy)):
        ws.cell(row=row,column=i+16).value=dummy[i]

def usage():
    print ("usage: " + sys.argv[0] + " -i <in.csv> -o <out.csv> [-c <col1>]", file=sys.stderr)
    print ("    Read a CSV, one (default last) column of which is a URL, then append a ", file=sys.stderr)
    print ("    bunch of Web derived values to the row and write out to the output file", file=sys.stderr)
    sys.exit(1)

#~
#if args.infile is None:
    #usage()

#if args.outfile is None:
    #usage()

#col=-1
#if args.col:
    #col=int(args.col)


count=0

# encoder options
jsonpickle.set_encoder_options('json', sort_keys=True, indent=2)
jsonpickle.set_encoder_options('simplejson', sort_keys=True, indent=2)

# default timeout for zgrab, in seconds
ztimeout='--timeout 15'#'--timeout 4'
# port parameters
pparms={ 
        '80': '-port 80 --lookup-domain',#added --lookup-domain
        '443-with-sni': '--port 443 --lookup-domain --tls',
        '443': '-port 443 -tls',
        }

r=[ws['B'],ws['F'],ws['G']] # 3 urls
numDifs=0 # number of schools with differences between 3 given urls

for row in range(len(r[0])):
#for row in range(10):
    
    urls=[]
    for col in r: # get all unique urls for this school
        if col[row].value not in urls and col[row].value: # only add unique
            #remove .ie/.com/whatever?? url.ie and url.com are the same?
            urls.append(col[row].value)
    print('num unique urls:',len(urls))
    if len(urls) > 1: # just info that might be interesting: number of schools with differences between 3 given urls
        numDifs+=1
            
    # sorting urls by prioritising those with paths (e.g. eoiniosagain.ie/eoin over eoiniosagain.ie)
    paths=[]
    hosts=[]
    for url in urls: # prioritise urls with paths
        if url and '/' in url:
            paths.append(url)
            hosts.append(url.split('/')[1]) # if url with path doesn't work, try without path
        #elif not url?
        else:
            hosts.append(url)
    urls=paths
    urls.extend(hosts)

    statusCodes=[list(),list(),list()] # status codes found at each port (lists are ports)

    # add dummy row if no urls found for school
    if len(urls) == 0:
        print ("Row " + str(row) + " has no URLs")
        # add dummy cols to row,
        if len(statusCodes) == 0: # only add dummy col if there's no other option
            wr_dummy(row)
    
    # for each of 3 urls for school, trying to find best one
    for url in urls:
        #keep looping until find first with ip that doesnt return an error, or certain status code
        cont=False
    
        #should add these to analysis
        error=''
        redirect=''
        
        analysis={}
        if url is None or url=='':
            print ("Row " + str(row) + " has an empty URL")
            # add dummy cols to row,
            if len(statusCodes) == 0: # only add dummy col if there's no other option
                wr_dummy(row)
        else:
            print ("Doing " + url,'Row:',row)
            # figure hostname from url
            # and figure pathname from url
            hostend=url.find("/")
            if hostend==-1:
                host=url
                path='/'
            else:
                host=url[:hostend]
                path=url[hostend:]
            # figure IP addr from hostname #~do elsewhere?
            try:
                answer = myResolver.query(host, "A") 
                for rdata in answer: 
                    #pick last one
                    addr=rdata
            except Exception as e: # if no ip at url
                print('no ip found at',url)
                for port in statusCodes:
                    if len(port) == 0: # only add dummy col if there's no other option
                        wr_dummy(row)
                        break
                #print (sys.argv[0] + " DNS exception:" + str(e) + "host: " + host , file=sys.stderr)
                # we're done with this host...
                continue#break
            #print "addr: " + str(addr)
            analysis['ip']=str(addr)
            analysis['host']=host
            # place to accumulate values for csv row
            csvals=[]
            # do HTTP, then HTTP and TLS with SNI, then HTTP and TLS without SNI
            #order
            ports=["80", "443", "443-with-sni"]
            for i in range(len(ports)):
                port=ports[i]
                analysis[port]={}
                # fill in what we can or a dummy value so csv col count is ok
                # even if we hit an exception
                analysis[port]['host']=host
                analysis[port]['ip']=str(addr)
                analysis[port]['status_code']=-1
                analysis[port]['body_len']=-1
                analysis[port]['names']=''
                analysis[port]['fp']=''
                analysis[port]['timely']=-1
                analysis[port]['self_signed']=-1
                analysis[port]['rsalen']=-1
                analysis[port]['cipher_suite']=-1
                analysis[port]['browser_trusted']=-1

                try:
                    redirecting=True
                    while redirecting:
                        redirecting=False
                        reserr=False # set if we don't get an HTTP response
                        cmd='zgrab '+  pparms[port] + " -http " + path + ' ' + ztimeout
                        print ("cmd: " + cmd)
                        proc=subprocess.Popen(cmd.split(),stdin=subprocess.PIPE,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
                        # 443 is different due to SNI, need to provide host, not address
                        if port=='443-with-sni' or port=='80':
                            #print ("doing: port: " + port + " |" + host + "| " + cmd)
                            # man it took me a while to figure that "\n" below was needed;-(
                            pc=proc.communicate(input=(host + "\n").encode())
                        else:
                            #print ("doing: port: " + port + " |" + str(addr) + "| " + cmd)
                            pc=proc.communicate(input=str(addr).encode())
                    
                        lines=pc[0].decode().split('\n')
                        jinfo=json.loads(lines[1])
                        jres=json.loads(lines[0])

                        #account for data being empty
                        if 'response' in jres['data']['http']:
                            if 'body' in jres['data']['http']['response']:
                                body=jres['data']['http']['response']['body']
                            else:
                                #print('no body')
                                body="Unknown"
                            analysis[port]['status_code']=jres['data']['http']['response']['status_code']
                            if '30' in str(analysis[port]['status_code']):
                                if '://' in str(jres['error']): # if redirecting (and not some other weird error
                                    redirect+=port+'|'+str(jres['error'].split('Get ')[1].split(': s')[0]) + ' '
                                    #check that redirect instead
                                    if url not in str(jres['error'].split('://')[1].split(': s')[0]) and \
                                       ports.index(port) == 0: # only redirect from p80 (don't want different urls for different ports)
                                        
                                        url=str(jres['error'].split('://')[1].split(': s')[0])
                                        if 'www.' in url:
                                            url=url.split('www.')[1]
                                        print('redirecting to',url)
                                        redirecting=True
                                else:
                                    print('Error1:',jres['error'])
                                    error+=port+'|' + jres['error']+'$'#~weird delimeter
                        elif 'error' in jres:
                            #print ("Error1 doing " + str(addr) + ":"+ port + " " + cmd, file=sys.stderr)
                            print('Error1:',jres['error'])
                            body=jres['error']
                            error+=port + '|' + jres['error']+'$'
                            analysis[port]['status_code']=999
                            reserr=True
                            cont=True#~
                            #if len(statusCodes[i]) > 0:~~~~~~~
                            #    continue # skip
                        #elif no body, print that (why?)
                        else:
                            print ("Unknown error doing " + str(addr) + ":"+ port + " " + cmd, file=sys.stderr)
                            print (jres, file=sys.stderr)
                            print (jinfo, file=sys.stderr)
                            body="Unknown"
                            error+='Unknown\n'
                            analysis[port]['status_code']=666
                            reserr=True
                            #if len(statusCodes[port]) > 0:~~~~~~~
                            #    continue # skip
                    # put host in there, as it's not by default if we didn't use it
                    jres['host']=host
                    jres['ip']=str(addr)
                    # also put it into port specifics, as we'll flatten that in a 'mo
                    if port=='80':
                        # do any port 80 specifics
                        pass
                    elif reserr == False and (port=='443-with-sni' or port=='443'):
                        # do tls specifics
                        if 'response' in jres['data']['http']:
                            th=jres['data']['http']['response']['request']['tls_handshake']
                            fp=th['server_certificates']['certificate']['parsed']['subject_key_info']['fingerprint_sha256'] 
                            cert=th['server_certificates']['certificate']
                            analysis[port]['fp']=fp
                            #print jsonpickle.encode(th)
                            # get_tls is from the surveys repo
                            get_tls('FreshGrab.py',port,th,jres['ip'],analysis[port],scandate)#~
                            #print analysis[port]
                            names={}
                            get_certnames(port,cert,names)#~
                            # flatten out them there names
                            flatnames=""
                            for k in names:
                                flatnames += str(names[k])
                                flatnames += ";"
                            analysis[port]['names']=flatnames
                        elif 'error' in jres:
                            #print ("Error2 doing " + str(addr) + ":"+ port + " " + cmd, file=sys.stderr)
                            print('Error2:',jres['error'])
                            body=jres['error']
                            error+=port + ' ' + jres['error']+'\n'
                            print (body)
                            analysis[port]['status_code']=999
                            cont=True#?
                            if len(statusCodes[port]) > 0:#~~
                                continue # skip
                        else:
                            body="Unknown"
                            error+='Unknown\n'
                            analysis[port]['status_code']=666
                            if len(statusCodes[port]) > 0:#~~
                                continue # skip
                    analysis[port]['body_len']=len(body)
                    
                except Exception as e:
                    print(traceback.format_exc())
                    print (sys.argv[0] + " exception: " + str(e) + "host: " + host + " port: " + port , file=sys.stderr)
                    #print ("pc: " + str(pc))
                    sys.exit(1)
                # accumulate columns for csv
                # basic info for all ports
                csvals.append(analysis[port]['status_code'])
                csvals.append(analysis[port]['body_len'])
                if port != "80":
                    # add TLS stuff
                    csvals.append(analysis[port]['names'])
                    csvals.append(analysis[port]['fp'])
                    csvals.append(analysis[port]['timely'])
                    csvals.append(analysis[port]['self_signed'])
                    csvals.append(analysis[port]['rsalen'])
                    csvals.append(analysis[port]['cipher_suite'])
                    csvals.append(analysis[port]['browser_trusted'])

            # write to file
            #-will keep overwriting if it finds a better url
            for i in range(len(csvals)):
                ws.cell(row=row+1,column=i+16).value=csvals[i]
            ws.cell(row=row+1,column=38).value=error
            ws.cell(row=row+1,column=39).value=url
            ws.cell(row=row+1,column=40).value=redirect

        if not cont: # break if 20x status codes found
            break
    if len(urls) > 1:
        print('used',url)
    #ws.cell(row=row+1,column=8).value=url
    #ws.cell(row=row+1,column=9).value=addr
    if count % 10 == 0:
            print ("Did",str(count),"last:", urls, file=sys.stderr)
    count += 1

print('num differences in urls given:',numDifs)#note this includes empty urls...
wb.save('../schools5.xlsx')
