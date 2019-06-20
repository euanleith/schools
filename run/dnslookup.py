# 6
# gets and stores dns info

# cleans up email by removing abnormalities
#just checks for / as that's only relevant one here
def cleanUrl(url):
    if url:
        init=url
        if '/' in url:
            temp = url.split('/')
            for i in range(len(temp)): # remove any / at either end
                if not temp[i]:
                    url=''
                    for j in range(len(temp)):
                        if j != i:
                            url+=temp[j]
        if init != url:
            print(init,'->',url)
    return url 

import dns
from dns import resolver
from openpyxl import load_workbook

wb = load_workbook("../schools5.xlsx")
ws = wb.active

urls=ws['AM']
recordtypes=['AAAA', 'MX', 'DS', 'CAA', 'SPF', 'TXT','CNAME','A']
cols = [10,11,12,13,14,15,36,37]

emailCols = [ws['A'],ws['G']]
badUrls = {"eircom","gmail.com","hotmail.com","iol.ie","outlook.ie"}

for row in range(len(urls)):
    if urls[row].value:      
        for record in range(len(recordtypes)):
            u=[urls[row].value]
            
            if recordtypes[record] == 'MX': # note only checking MX
                #if not badUrl in email, make arr and check each of them
                for emails in emailCols:
                    if emails[row].value:
                        b=True
                        for bad in badUrls:
                            if bad in emails[row].value:
                                b=False
                                break
                        if b:
                            if '@' in emails[row].value:
                                email=emails[row].value.split('@')[1]
                            elif len(emails[row].value.split('.')) > 2:
                                email=emails[row].value.split('.')
                                email=email[len(email)-1]
                            email = cleanUrl(email)
                            if email not in u:
                                u.append(email)
                    
            for url in u:
                try:
                    hostend=url.find('/')
                    if hostend == -1:
                        host=url
                    else:
                        host=url[:hostend]
                    if recordtypes[record] == 'CNAME':
                        host='www.'+host
                    ans = dns.resolver.query(host, recordtypes[record])
                    # some have multiple, so add all, seperated by spaces
                    txt=""
                    for answer in ans.response.answer:
                        for item in answer.items:
                            txt += item.to_text() + " "
                    ws.cell(row=row+1,column=cols[record]).value=txt
                    if u.index(url) > 0:
                        print('-',recordtypes[record],'at',url)
                    break
                except dns.resolver.NoAnswer:
                    # if no result for given record type
                    ws.cell(row=row+1,column=cols[record]).value="no " + recordtypes[record]
                    # printing if find something unusual
                    if (recordtypes[record] == 'MX' or recordtypes[record] == 'TXT' \
                        or recordtypes[record] == 'A'):
                        print("no " + recordtypes[record] + " at " + url)
                except dns.resolver.NXDOMAIN:
                    # if invalid domain
                    # checked in cmpurls, but here too just to be sure
                    if not ws.cell(row=row+1,column=cols[record]).value:
                        ws.cell(row=row+1,column=cols[record]).value=""
                    print("Invalid domain:",url)
                    continue
                except Exception as e:
                    print(e)
                    if not ws.cell(row=row+1,column=cols[record]).value:
                        ws.cell(row=row+1,column=cols[record]).value=""
                    continue
                else:
                    # printing if find something unusual
                    if (recordtypes[record] != 'MX' and recordtypes[record] != 'TXT' and \
                        recordtypes[record] != 'A'):
                        print(recordtypes[record] + " at " + url)

wb.save("../schools2019.xlsx")
