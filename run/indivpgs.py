# 2
# given indiv pgs, get emails and web-addresses
# so have to go into given domain name's html, search for prefixes
#'email address:' and 'website address:' and store that info
#needs testing

from requests_html import HTMLSession # note: should use most recent version?
from openpyxl import load_workbook

session = HTMLSession()

wb = load_workbook("../schools.xlsx")
ws = wb.active
rolls = ws['D']


# string prefixes are messy but works
for x in range(len(rolls)):
#for x in range(10):
    try:
        print("doing " + rolls[x].value.split('roll=')[1])
        r = session.get(rolls[x].value)

        # store emails
        email = r.html.search('mailto:{}">')
        if email:
            email = email[0]
            if "@" in email:
                emailfound = email.split("@")[1]
            else:
                print("Unkown email: " + email)
                emailfound = ""
        else:
            print("no email found")
            emailfound = ""
            
        ws.cell(row=x+1,column=6).value=emailfound

            
        # store urls
        #some are email addresses... e.g. moatecs@eircom.com
        url = r.html.search('Website address:</strong><span><a target="_blank" href="{}">')
        if url:
            url = url[0]
            if "http://www." in url:
                urlfound = url.split("http://www.")[1]
            elif "http://" in url:
                urlfound = url.split("http://")[1]
            else:
                urlfound=""
                print("Unknown url: " + url)
            #if "@" in urlfound: # some are emails...
            #    urlfound = urlfound.split("@")[1]
        else:
            print("no url found")
            urlfound = ""         
        
        ws.cell(row=x+1,column=7).value=urlfound
        
    except Exception as e:
        print("Error:", e)
    #list index out of range


wb.save("../schools2.xlsx")
