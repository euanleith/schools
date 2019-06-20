# 3
# clean list
#   remove eircom.net etc.
#       eircom.net, gmail, hotmail, yahoo, other...
#       check recurring for recognisable names
#   remove duplicates
# clean up urls

from openpyxl import Workbook
from openpyxl import load_workbook
import operator

# adds to num duplicates of txt
# (as dict of {name:num})
def addDups(txt, dct):
    if txt in dct:
        dct[txt] += 1
    else:
        dct.update({txt:1})
    return dct

def orderDict(dct,minDup):
    temp=dict()
    for key in dct:
        if dct[key] > minDup:
            temp.update({key:dct[key]})
    return dict(sorted(temp.items(), key=lambda item: item[1]))

# cleans up email by removing abnormalities
def cleanUrl(url):
    if url:
        init=url
        if 'www' in url:
            if 'www.' in url:
                temp = url.split('www.')
            else:
                temp = url.split('www')
            url=temp[1]
        if 'http' in url:
            if 'http//' in url:
                temp = url.split('http//')
            else:
                temp = url.split('http')
            url=temp[1]
        if len(url.split('.')) > 2:
            temp = url.split('.')
            if not temp[0]:
                url=''
                for i in range(1,len(temp)):
                    url+=temp[i]
                    if i < len(temp)-1:
                        url+='.'
            else:
                print("couldn't fix:",url)
        if not '.' in url:
            url += '.com' #not great, but can't really test each of .com/.i.e/etc.
        if ' ' in url:
            temp = url.split(' ')
            url=''
            for t in temp:
                url+=t
        if '/' in url:
            temp = url.split('/')
            for i in range(len(temp)): # remove any / at either end
                if not temp[i]:
                    url=''
                    for j in range(len(temp)):
                        if j != i:
                            url+=temp[j]
        if '@' in url:
            temp = url.split('@')
            b=False
            for u in badUrls: # if badUrl in url, take other part
                if u in temp[1]:
                    b=True
                    break
            if b:
                url = temp[0]+'.ie'
            else:
                url = temp[1]
        if init != url:
            print(init,'->',url)
    return url 


badUrls = {"eircom","gmail.com","hotmail.com","iol.ie","outlook.ie",\
           "lcetb.ie","wwetb.ie","msletb.ie","kwetb.ie","lwetb.ie","ddletb.ie",\
           "donegaletb.ie","lmetb.ie","gretb.ie","cdetb.ie"}
#remove lcetb etc., but still want to check them...


wb = load_workbook("../schools2.xlsx")
ws = wb.active

cols=[0x42,0x46] # B, F

print('Cleaning emails:')
for col in cols:
    dups=dict()
    emails=ws[chr(col)]
    for i in range(len(emails)):
        email=emails[i].value
        if email:
            dup=False
            for url in dups: # check if dup
                if url in email:
                    ws.cell(row=i+1,column=col-0x40).value="" # remove duplicates
                    addDups(url,dups)
                    dup=True
                    break
            if not dup:
                bad=False
                for url in badUrls: # check if 'bad'
                    if url in email:
                        ws.cell(row=i+1,column=col-0x40).value="" # remove 'bad urls'
                        addDups(url,dups)
                        bad=True
                        break
                if not bad:
                    addDups(email,dups)



    dups=orderDict(dups,2)
    print('\n['+chr(col)+']:')
    print("Total num duplicates: " + str(sum(dups.values())) + " from " + str(len(dups)) + " hosts")
    print("Total num uniques: " + str(len(emails)-sum(dups.values())))
    for key in dups:
        print("%s: %s" % (key, dups[key]))

urls= ws['G']
print('\nCleaning urls:\n')
for row in range(len(urls)):
    url=cleanUrl(urls[row].value)
    ws.cell(row=row+1,column=7).value=url

wb.save("schools3.xlsx")


# new workbook for dups
wb = Workbook()
ws = wb.active

row=1
for key in dups:
    if 'etb' in key: #donegalvec? links to donegaletb...
        ws.cell(row=row+1,column=1).value=key
        ws.cell(row=row+1,column=2).value=dups[key]
        row+=1

wb.save('../schools3.xlsx')
        
