# 1
# get useful data from initial spreadsheet
#weird email case: user.site.ie

from openpyxl import load_workbook
from openpyxl import Workbook


# load data

wb = load_workbook("../list.xlsx")
ws = wb.active      # ws = worksheet with data

emails = ws['L']     # get emails cell 'L'
emails = emails[2:]   # remove title emails

rolls = ws['A']
rolls = rolls[2:len(rolls)-3]

eircodes = ws['I']
eircodes = eircodes[2:]



# store data

newwb = Workbook()  # workbook to store data
newws = newwb.active
indivPrefix = "https://www.education.ie/en/find-a-school/School-Detail/?roll="; # prefix for finding individual school's page on education.ie

#might just be func for putting arr items in col...

for x in range(len(emails)):# do for collength instead (more general)
    # store email
    newws.cell(row=x+1,column=1).value=emails[x].value   

    # store domain name
    if (emails[x].value) and ("@" in emails[x].value):
            newws.cell(row=x+1,column=2).value=emails[x].value.split("@")[1]

    # store roll number
    if x < len(rolls): # shouldnt be necessary...
        newws.cell(row=x+1,column=3).value=rolls[x].value

    # store inidividual pgs from education.ie
    if x < len(rolls):
        newws.cell(row=x+1,column=4).value=indivPrefix+rolls[x].value

    #store eircodes
    newws.cell(row=x+1,column=5).value=eircodes[x].value
    
newwb.save("../schools.xlsx")
