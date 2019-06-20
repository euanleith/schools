# 4
# check for duplicate urls in data
# if so, check each to see if there's an ip associated with one of the urls, and take that one
# else, check if there's an ip associated with the common url
# stores these ipv4 ips
#note: socket.getaddrinfo for more detail
#use socket.getfqdn(url) to get host (e.g. squarespace)?
#can use zmap to get ips
#needs testing

from openpyxl import load_workbook
import socket

wb = load_workbook("../schools3.xlsx")
ws = wb.active

# cmp cols B, F, G

urls1 = ws['B']
urls2 = ws['F']
urls3 = ws['G']
rolls = ws['C']
urls = {urls1, urls2, urls3}


for x in range(len(urls1)):
#for x in range(10):
    # check if dif
    #ignore if .ie and .com?
    if (urls1[x].value and urls2[x].value and urls1[x].value.split(".")[0] != urls2[x].value.split(".")[0]) or \
    (urls1[x].value and urls3[x].value and urls1[x].value.split(".")[0] != urls3[x].value.split(".")[0]) or \
    (urls2[x].value and urls3[x].value and urls2[x].value.split(".")[0] != urls3[x].value.split(".")[0]):
        txt="dif at " + str(x) + ": "

        #if ones NoneType, dont type it
        #if the only difference between them is NoneType
        if urls1[x].value:
              txt += urls1[x].value + " "
        if urls2[x].value:
              txt += urls2[x].value + " "
        if urls3[x].value:
              txt += urls3[x].value + " "
        print(txt)
        # check each url to see if ip associated
        txt = "no ip at any urls"
        ips=[]
        for url in urls:#should be for unique urls (set)
            if url[x].value:
                try:
                    hostend=url[x].value.find('/')
                    if hostend == -1:
                        host=url[x].value
                    else:
                        host=url[x].value[:hostend]
                    ip = socket.gethostbyname(host)
                except socket.error as e: # if no ip at url
                    continue
                else: # if ip at url, store
                    #print("ip at " + url[x].value + ": " + ip)
                    ips.append(ip)
                    txt=None
                    ws.cell(row=x+1,column=8).value=url[x].value
                    ws.cell(row=x+1,column=9).value=ip
                    if hostend != -1:
                        break; # break if host with path and ip associated, else just takes last
        # if no ips found            
        if txt:
            print(txt)
        # if dif ips found
        setips=list(set(ips))
        if (len(setips) > 1):
            txt = "dif ips: "
            for i in range(len(setips)):
                txt += setips[i] + " "
            print(txt)
        else:
            print("one ip at",ip)
    else: #if all same, just add regardless of if ip found?
        if urls3[x].value:
            #print(socket.getaddrinfo(urls3[x].value, 80));
                  
            try:
                hostend=urls3[x].value.find('/')
                if hostend == -1:
                    host=urls3[x].value
                else:
                    host=urls3[x].value[:hostend]
                ip = socket.gethostbyname(host)
            except socket.error as e: # if no ip at url
                print("no ip at", urls3[x].value)
            else: # if ip at url, store
                ws.cell(row=x+1,column=8).value=urls3[x].value
                ws.cell(row=x+1,column=9).value=ip 
                #print(ip)
        else:
            print("no urls for",rolls[x].value)


wb.save("../schools4.xlsx")
