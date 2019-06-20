from openpyxl import load_workbook

wb = load_workbook("../schools2019.xlsx")
ws = wb.active

cols=[ws['A'],ws['F'],ws['G']]
names=['emails from spreadsheet','emails from education.ie','urls from education.ie']
none=[[],[],[]]
rolls=ws['C']

for col in range(len(cols)):
    for row in range(len(cols[col])):
        if not cols[col][row].value:
            none[col].append(rolls[row].value)

for col in range(len(none)):
    print(names[col]+':',len(none[col]))
