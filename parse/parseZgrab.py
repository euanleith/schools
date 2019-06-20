from openpyxl import load_workbook

import sys,os
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import funcs

ports=['443 without sni','443 with sni','80'] # note order
infoNames=["status codes","names","fp","cipher suites","key lengths",\
               "not timely","self signed","browser not trusted"]
groups=['all','bools','dups']


out=dict()

years=["schools2018.xlsx","schools2019.xlsx"]
for year in years:
    wb = load_workbook('../'+year)
    ws = wb.active
    year=year.split('schools')[1].split('.')[0]

    # columns for each zgrab thing, for each port
    statusCodes=[ws['R'],ws['AA'],ws['P']]
    names=[ws['T'],ws['AC']]
    fp=[ws['U'],ws['AD']]
    cipherSuites=[ws['Y'],ws['AH']]
    keyLengths=[ws['X'],ws['AG']] # (rsalen)

    timely=[ws['V'],ws['AE']]
    selfSigned=[ws['W'],ws['AF']]
    browserTrusted=[ws['Z'],ws['AI']]

    urls=ws['H']


    # columns to check
    dupInfo=[statusCodes,names,fp,cipherSuites,keyLengths]#search for all schools with a given dupInfo thing, e.g. all with 404 in statusCodes
    boolInfo=[timely,selfSigned,browserTrusted]

    out.update({year:{ports[0]:dict(),ports[1]:dict(),ports[2]:dict()}})

    # for each port
    for port in range(len(ports)):
        
        # for each zgrab thing (looking for dups) (status codes, names) (returns dict)
        for j in range(len(dupInfo)):
            info={} # dict for number of instances of each element in a column

            #remove dud ones (with -1 in statusCodes)
            # for each row
            if port < len(dupInfo[j]):
                for row in dupInfo[j][port]:
                    funcs.addDups(info,str(row.value))

                out[year][ports[port]].update({infoNames[j]:info})

        # for each zgrab boolean (not timely, self signed, browser not trusted) (returns num negatives)
        for j in range(len(boolInfo)):
            if port < len(boolInfo[j]):
                true=[]
                false=[]
                for row in range(len(boolInfo[j][port])):
                    if urls[row].value and str(boolInfo[j][port][row].value):
                        if str(boolInfo[j][port][row].value).lower() == 'false':
                            false.append(urls[row].value)
                        elif str(boolInfo[j][port][row].value).lower() == 'true':
                            true.append(urls[row].value)
                out[year][ports[port]].update({infoNames[j+len(dupInfo)]:[true,false]})

# user input
while True:
    try:
        inp=input("port/info: ")
        if inp == '?':
            print()
            print("ports:",ports)
            print("info:",infoNames)
            print("groups:",groups)
        elif inp == 'q':
            break;

        # if port chosen
        elif inp in ports:
            port=inp
            while True:    
                info=input("info: ")
                if info == '?':
                    print(infoNames[:len(out['2018'][port])],'\n')#~
                elif info == 'q':
                    break;
                else:
                    if info in infoNames[:len(dupInfo)]: # if dup info
                        # order and get minDups if relevant
                        numDups=int(input('num duplicates: '))
                        for year in out:
                            t=out[year][port][info]
                            t=funcs.sortDict(t,numDups)
                            print('\n'+year+':',t)
                            #print('num results: ' + str(len(t)))#/total
                            print('num results:',str(sum(funcs.sortDict(t,numDups).values())))
                    else: # else if bool info
                        for year in out:
                            print(year)
                            true=out[year][port][info][0]
                            false=out[year][port][info][1]
                            if info == 'self signed':
                                print(true)
                                print(info+':',len(true),', not ' + info + ':',len(false))
                            else:
                                print(false)
                                print(info.split('not ')[1] + ':',len(true),\
                                      ', ' + info + ':',len(false))
                print()
                        
        # else if going straight to thing (do all relevant ports)
        elif inp in infoNames:
            info=inp
            if info in infoNames[:len(dupInfo)]: # if dup info
                    minDups = int(input('min duplicates: '))
            for port in ports:
                if info in out['2018'][port]:#~
                    print('\n['+port+']:')
                for year in out:
                    if info in out[year][port]:
                        if info in infoNames[:len(dupInfo)]: # if dup info
                            t=out[year][port][info]
                            #print(year + ': ' + str(len(t)))
                            #print(year+':',str(sum(funcs.sortDict(t,minDups).values())))
                            print(year + ':',funcs.sortDict(t,minDups))#hard to see...
                        else: # else if bool info
                            true=out[year][port][info][0]
                            false=out[year][port][info][1]
                            if info == 'self signed':
                                print(year+':',info+':',len(true),', not ' + info + ':',len(false))
                            else:
                                print(year+':',info.split('not ')[1] + ':',len(true),\
                                      ', ' + info + ':',len(false))
                        #might want to print more?

        # else if 'all'       
        elif inp == 'all' or inp == 'bools' or inp == 'dups':#dups is useless
            if inp == 'dups':
                minDups=int(input('min duplicates: '))
            for port in ports:
                if not (inp == 'bools' and port == '80'):#messy
                    print('\n['+port+']:')
                    for year in out:
                        print('\n'+year+':')
                        for info in out[year][port]:
                            if info in infoNames[:len(dupInfo)]:
                                if inp != 'bools': # if dup info
                                    t=out[year][port][info]
                                    print(info + ': ' + str(sum(funcs.sort(t,minDups).values())))
                                    #print(info + ':',t)#~
                            elif inp != 'dups': # else if bool info
                                true=out[year][port][info][0]
                                false=out[year][port][info][1]
                                if info == 'self signed':
                                    print(info+':','{0:.3f}'.format(len(true)/(len(false)+len(true))),\
                                          ', not ' + info + ':','{0:.3f}'.format(len(false)/(len(false)+len(true))))
                                else:
                                    print(info.split('not ')[1] + ':','{0:.3f}'.format(len(true)/(len(false)+len(true))),\
                                          ', ' + info + ':','{0:.3f}'.format(len(false)/(len(false)+len(true))))
                                    #either;
                                    #'{0:.3f}'.format(len(true or false)/(len(false)+len(true)))
                                    #len(true or false)
        else:
            print("invalid input:",inp)
        print()
    except Exception as e:
        print("invalid input:",e,'\n')
        continue
    
    
