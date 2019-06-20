from openpyxl import load_workbook

import sys,os
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import funcs

years=["schools2018.xlsx","schools2019.xlsx"]
infos=[['K'],['T','AC']]#should do both names
infoNames=['MX','Names']
colNames=[[''],['443 without sni','443 with sni']]
for i in infos:
    print('\n'+infoNames[infos.index(i)]+':')
    for col in i:
        if colNames[infos.index(i)][i.index(col)]:
            print('\n'+colNames[infos.index(i)][i.index(col)]+':')
        for year in years:
            wb = load_workbook('../'+year)
            ws = wb.active
            year=year.split('schools')[1].split('.')[0]
            print('\n['+year+']:')

            own=[]
            other=dict()

            urls=ws['H']
            info=ws[col]
                
            hosts={'google':0,'outlook':0,'blacknight':0,'hosts':0,\
               'reg365':0,'irishdomains':0,'cloudaccess':0,\
               'tstechnology':0,'mailspamprotection':0,'elive':0,\
               'host':0,'scoilnet':0,'spamtitan':0,'spamdoctor':0,\
               'forwardmx':0,'messagelabs':0,'novara':0,\
                'odin':0,'plesk':0,'weebly':0,'parallels panel':0,\
                'squarespace':0,'avenir':0,'carrigtwohill':0}
            okhosts={'wwetb':0, 'msletb':0, 'kwetb':0, 'lwetb':0, 'ddletb':0,\
                     'donegaletb':0, 'lmetb':0, 'gretb':0, 'lcetb':0,\
                     'donegalvec':0, 'tipperaryetb':0, 'corketb':0,\
                     'wicklowvec':0, 'kilkennycarlowetb':0}
            #~maybe not exhaustive/not all should be there
            
            for row in range(len(info)):
                if info[row].value and info[row].value != 'no MX':
                    hosted=False
                    for host in hosts:
                        if host in info[row].value.lower():
                            hosts[host]+=1
                            hosted=True
                    if not hosted:
                        for host in okhosts:
                            if host in info[row].value.lower():
                                okhosts[host]+=1
                                hosted=True
                    if not hosted:
                        if urls[row].value and urls[row].value.split('.')[0] not in info[row].value:
                            other=funcs.addDups(other,info[row].value)
                        else:
                            own.append(info[row].value)
                                
            numHosted=0
            for host in hosts:
                numHosted+=hosts[host]
            numOther=0
            for o in other:
                numOther+=other[o]
            numOkHosted=0
            for host in okhosts:
                numOkHosted+=okhosts[host]

            print('-num self owned:',len(own),\
                  '('+'{0:.3f}'.format(len(own)/(len(own)+numHosted+numOkHosted+numOther))+')')
            print('-num schools hosted:',numHosted,\
                  '('+'{0:.3f}'.format((numHosted)/(len(own)+numHosted+numOkHosted+numOther))+')')
            print('-num schools ok hosted:',numOkHosted,\
                  '('+'{0:.3f}'.format((numOkHosted)/(len(own)+numHosted+numOkHosted+numOther))+')')#phrase better
            print('-num hosts:',len(hosts)+len(other))

            #less important info
            print('-categorised hosts (>5):',funcs.sortDict(hosts,5))
            #print('-num categorised hosts:',len(hosts))
            #print('-ok hosts:',orderDict(okhosts,0))
            #print('-other hosts (>5):',orderDict(other,5))
            #print('-num other hosts:',len(other))
            
