from openpyxl import load_workbook

years=["schools2018.xlsx","schools2019.xlsx"]
ports=['443 without sni','443 with sni']
cols=['Y','AH']

for col in cols:
    print('\n['+ports[cols.index(col)]+']:')
    for year in years:
        wb = load_workbook('../'+year)
        ws = wb.active
        year=year.split('schools')[1].split('.')[0]
        #print('['+year+']:')

        #cipherSuites=[ws['Y'],ws['AH']]
        port=ws[col]
        urls=ws['H']

        #for port in cipherSuites:
        bad={0x000A:[],0x0035:[],0xC011:[],0x002F:[],0x0005:[]}#could add more
        numCipherSuites=0
        for cipherSuite in port:
            if cipherSuite.value:
                numCipherSuites+=1
                if cipherSuite.value in bad:
                    bad[cipherSuite.value].append(urls[port.index(cipherSuite)])
        #print(ports[cipherSuites.index(port)]+':')
        txt=''
        for b in bad:
            txt+=str(hex(b))+': '+str('{0:.3f}'.format(len(bad[b])/numCipherSuites))+', '
        print(year+':',txt)
        #print total proportion of bad cipher suites
print()
            
