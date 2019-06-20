from openpyxl import load_workbook

ports=['443 without sni','443 with sni','80'] # note order
infoNames=["status code","name","fp","cipher suite","key length",\
               "not timely","self signed","browser not trusted"]

#should also show dns stuff

while (True):
    inp=input('school: ')

    if inp == 'q':
        break

    years=["schools2018.xlsx","schools2019.xlsx"]
    for year in years:
        wb = load_workbook('../'+year)
        ws = wb.active
        year=year.split('schools')[1].split('.')[0]

        statusCodes=[ws['R'],ws['AA'],ws['P']]
        names=[ws['T'],ws['AC']]
        fp=[ws['U'],ws['AD']]
        cipherSuites=[ws['Y'],ws['AH']]
        keyLengths=[ws['X'],ws['AG']] # (rsalen)

        timely=[ws['V'],ws['AE']]
        selfSigned=[ws['W'],ws['AF']]
        browserTrusted=[ws['Z'],ws['AI']]

        urls=ws['H']

        infos=[statusCodes,names,fp,cipherSuites,keyLengths,timely,selfSigned,browserTrusted]

        schools=[]
        for url in urls:
            if url.value and inp in url.value:
                schools.append(url)
        if len(schools) > 1:
            print('multiple schools found:',[school.value for school in schools])
            break
        elif len(schools) == 0:
            print('['+year+']:')
            print('no schools found')
        
        for school in schools:
            print('\n'+school.value)
            print('['+year+']:')
            i = urls.index(school)
            for port in ports:
                print('\n'+port+':')
                for info in infos:
                    if ports.index(port) < len(info):
                        print(infoNames[infos.index(info)]+':',info[ports.index(port)][i].value)
