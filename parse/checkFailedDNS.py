# looks through the urls for which dnslookup didn't work
from openpyxl import load_workbook

wb = load_workbook('../schools2019.xlsx')
ws = wb.active

urlsfound=ws['H']
urls=[ws['B'],ws['F'],ws['G']]

for row in range(len(urlsfound)):
    if not urlsfound[row].value:
        for url in urls:
            if url[row].value:
                print(url[row].value)
            else:
                print('-')
        print()
    
