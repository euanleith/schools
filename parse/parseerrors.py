from openpyxl import load_workbook

import sys,os
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import funcs

wb = load_workbook('../schools2019.xlsx')
ws = wb.active

errors=ws['AL']
errortypes={'oversized record received':0,'connection reset by peer':0,\
            'internal error':0,'handshake failure':0,'i/o timeout':0,\
            'EOF':0,'connection refused':0,'stopped after 0 redirects':0,\
            'network is unreachable':0}

for row in errors:
    if row.value:
        error=row.value.split('\n')
        for e in error:
            for errortype in list(errortypes.keys()):
                if errortype in e:
                    funcs.addDups(errortypes,errortype)
                    break
errortypes=funcs.sortDict(errortypes)
print(errortypes)
