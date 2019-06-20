from openpyxl import load_workbook

ports=['443 without sni','443 with sni']

years=["schools2018.xlsx","schools2019.xlsx"]
for year in years:
    wb = load_workbook('../'+year)
    ws = wb.active
    year=year.split('schools')[1].split('.')[0]
    print('\n['+year+']:')

    keyLengths=[ws['X'],ws['AG']]
    urls=ws['H']

    for port in keyLengths:
        txt=ports[keyLengths.index(port)]+':'
        cnt=0
        for row in range(len(port)):
            val=port[row].value
            if val and bin(int(val)).count('1') != 1: # or too small?
                txt += url[row] + ', '
                cnt+=1
        if cnt == 0:
            txt += ' all ok'
        print(txt)
